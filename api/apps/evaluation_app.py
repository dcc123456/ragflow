#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
RAG Evaluation API Endpoints

Provides REST API for RAG evaluation functionality including:
- Collection management
- Test case management
- Evaluation execution
- Results retrieval
- Configuration recommendations
"""

import json
import logging
import re
import unicodedata
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from quart import request, send_file
from api.apps import login_required, current_user
from api.db.db_models import DB, EvaluationCase, EvaluationResult
from api.db.services.evaluation_service import EvaluationService
from api.db.services.task_service import TaskService
from api.db.services.user_service import UserTenantService
from api.utils.api_utils import (
    get_data_error_result,
    get_json_result,
    get_request_json,
    server_error_response,
    validate_request
)
from common.constants import RetCode
from api.common.priority_provider import get_tenant_priority
from deepdoc.parser.excel_parser import RAGFlowExcelParser
from rag.utils.redis_conn import REDIS_CONN


def _has_collection_access(collection_id: str) -> bool:
    """Check if current user has access to an evaluation collection.

    Access is granted if the user is the owner (tenant_id match) or a
    team member of the tenant that owns the collection.
    """
    collections = EvaluationService.query(tenant_id=current_user.id, id=collection_id)
    if collections:
        return True
    ok, collection = EvaluationService.get_by_id(collection_id)
    if not ok or not collection:
        return False
    return bool(UserTenantService.filter_by_tenant_and_user_id(
        collection.tenant_id, current_user.id))


_XLSX_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_XLSX_CELL_MAX_LEN = 32767


def _sanitize_xlsx_filename(name: str) -> str:
    name = unicodedata.normalize("NFKC", str(name or "")).strip()
    name = re.sub(r"[\x00-\x1f\x7f]", " ", name)
    name = re.sub(r'[<>:"/\\\\|?*]+', " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = "collection_cases"
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name[:200]


def _xlsx_safe_text(value) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    if not value:
        return ""
    value = _XLSX_ILLEGAL_CHARS_RE.sub("", value)
    if len(value) > _XLSX_CELL_MAX_LEN:
        value = value[:_XLSX_CELL_MAX_LEN]
    return value


# ==================== Collection Management ====================

@manager.route('/collection/create', methods=['POST'])  # noqa: F821
@login_required
async def create_collection():
    """
    Create a new evaluation collection and import cases from an Excel file.

    Multipart form:
    - file: Excel/CSV file containing columns: question, reference_answer
    - name: Optional collection name (default: file name without extension)
    - description: Optional description
    """
    try:
        form = await request.form
        name = (form.get("name") or "").strip()
        description = form.get("description", "")
        target_type = (form.get("target_type") or "chat").strip().lower()
        files = await request.files
        file = files["file"] if files and files.get("file") else None
        if not file:
            return get_data_error_result(message="Excel file is required")

        if not name:
            filename = file.filename or ""
            name = Path(filename).stem.strip()

        if not name:
            return get_data_error_result(message="Collection name cannot be empty")

        file_bytes = file.read()
        try:
            workbook = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(file_bytes))
        except Exception as e:
            return get_data_error_result(message=f"Failed to parse Excel file: {e}")

        cases = []
        for sheetname in workbook.sheetnames:
            ws = workbook[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            headers = [
                (str(cell.value).strip() if cell.value is not None else "")
                for cell in rows[0]
            ]
            header_map = {h.lower(): idx for idx, h in enumerate(headers) if h}
            if "question" not in header_map or "reference_answer" not in header_map:
                return get_data_error_result(
                    message=f"Sheet '{sheetname}' must contain columns: question, reference_answer"
                )
            q_idx = header_map["question"]
            a_idx = header_map["reference_answer"]
            for row in rows[1:]:
                question = ""
                reference_answer = ""
                if q_idx < len(row) and row[q_idx].value is not None:
                    question = str(row[q_idx].value).strip()
                if a_idx < len(row) and row[a_idx].value is not None:
                    reference_answer = str(row[a_idx].value).strip()
                if not question and not reference_answer:
                    continue
                cases.append({
                    "question": question,
                    "reference_answer": reference_answer
                })

        if not cases:
            return get_data_error_result(message="No valid cases found in file")

        success, result = EvaluationService.create_collection(
            name=name,
            description=description,
            tenant_id=current_user.id,
            user_id=current_user.id,
            target_type=target_type,
        )

        if not success:
            return get_data_error_result(message=result)

        success_count, failure_count = EvaluationService.import_test_cases(
            collection_id=result,
            cases=cases
        )

        return get_json_result(data={
            "collection_id": result,
            "success_count": success_count,
            "failure_count": failure_count,
            "total": len(cases),
        })
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/list', methods=['GET'])  # noqa: F821
@login_required
async def list_collections():
    """
    List evaluation collections for current tenant.

    Query params:
    - page: Page number (default: 1)
    - page_size: Items per page (default: 20)
    - keywords: Optional search keywords for collection title
    """
    try:
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 20))
        keywords = (request.args.get("keywords", "") or "").strip()

        result = EvaluationService.list_collections(
            tenant_id=current_user.id,
            user_id=current_user.id,
            page=page,
            page_size=page_size,
            keywords=keywords,
        )

        return get_json_result(data=result)
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>', methods=['GET'])  # noqa: F821
@login_required
async def get_collection(collection_id):
    """Get collection details by ID"""
    try:
        if not EvaluationService.query(
                tenant_id=current_user.id, id=collection_id):
            return get_json_result(
                data=False, message='Collection not configured or access denied',
                code=RetCode.OPERATING_ERROR)
        success, collection = EvaluationService.get_by_id(collection_id)
        if not success:
            return get_data_error_result(
                message="Collection not found",
                code=RetCode.DATA_ERROR
            )

        return get_json_result(data=collection.to_dict())
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>', methods=['PUT'])  # noqa: F821
@login_required
async def update_collection(collection_id):
    """
    Update collection.

    Request body:
    {
        "name": "New name",
        "description": "New description",
    }
    """
    try:
        if not EvaluationService.query(
                tenant_id=current_user.id, id=collection_id):
            return get_json_result(
                data=False, message='Collection not configured or access denied',
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        req = {
            "name": req.get("name"),
            "description": req.get("description"),
        }

        success = EvaluationService.update_collection(collection_id, **req)
        if not success:
            return get_data_error_result(message="Failed to update collection")

        return get_json_result(data={"collection_id": collection_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>', methods=['DELETE'])  # noqa: F821
@login_required
async def delete_collection(collection_id):
    """Delete collection"""
    try:
        success = EvaluationService.delete_collection(collection_id)

        if not success:
            return get_data_error_result(message="Failed to delete collection")

        return get_json_result(data={"collection_id": collection_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/remove', methods=['POST'])  # noqa: F821
@login_required
async def delete_collections():
    """Delete collections"""
    req = await get_request_json()
    collection_ids = req.get("collection_ids", [])
    try:
        for collection_id in collection_ids:
            EvaluationService.delete_collection(collection_id)
        return get_json_result(data=True)
    except Exception as e:
        return server_error_response(e)

# ==================== Test Case Management ====================

@manager.route('/collection/<collection_id>/case/add', methods=['POST'])  # noqa: F821
@login_required
async def add_test_case(collection_id):
    """
    Add a test case to a collection.

    Request body:
    {
        "question": "Test question",
        "reference_answer": "Optional ground truth answer",
        "relevant_doc_ids": ["doc_id1", "doc_id2"],
        "relevant_kb_ids": ["kb_id1", "kb_id2"],
        "metadata": {"key": "value"}
    }
    """
    try:
        if not _has_collection_access(collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        question = req.get("question", "").strip()

        if not question:
            return get_data_error_result(message="Question cannot be empty")
        variable = {
            "question": question,
            "reference_answer": req.get("reference_answer", "").strip()
        }
        success, result = EvaluationService.add_test_case(
            collection_id=collection_id,
            variable=variable,
            relevant_doc_ids=req.get("relevant_doc_ids"),
            relevant_kb_ids=req.get("relevant_kb_ids"),
            metadata=req.get("metadata")
        )

        if not success:
            return get_data_error_result(message=result)

        return get_json_result(data={"case_id": result})
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>/case/import', methods=['POST'])  # noqa: F821
@login_required
@validate_request("cases")
async def import_test_cases(collection_id):
    """
    Bulk import test cases.

    Request body:
    {
        "cases": [
            {
                "question": "Question 1",
                "reference_answer": "Answer 1",
                ...
            },
            {
                "question": "Question 2",
                ...
            }
        ]
    }
    """
    try:
        if not _has_collection_access(collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        cases = req.get("cases", [])

        if not cases or not isinstance(cases, list):
            return get_data_error_result(message="cases must be a non-empty list")

        success_count, failure_count = EvaluationService.import_test_cases(
            collection_id=collection_id,
            cases=cases
        )

        return get_json_result(data={
            "success_count": success_count,
            "failure_count": failure_count,
            "total": len(cases)
        })
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>/case/export', methods=['GET'])  # noqa: F821
@login_required
async def export_test_cases(collection_id):
    """
    Export all test cases for a collection as an Excel file (.xlsx).

    Response:
    - Excel file stream (attachment)
    - Columns: id, question, reference_answer, relevant_doc_ids, relevant_kb_ids, metadata, create_time
    """
    try:
        if not _has_collection_access(collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)

        ok, collection = EvaluationService.get_by_id(collection_id)
        if not ok or not collection:
            return get_data_error_result(message="Collection not found")

        cases = EvaluationService.get_test_cases(collection_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "cases"
        ws.append([
            "id",
            "question",
            "reference_answer",
            "create_time",
        ])

        for case in cases:
            variable = case.get("variable") or {}
            ws.append([
                case.get("id", ""),
                variable.get("question", ""),
                variable.get("reference_answer", ""),
                case.get("create_time", ""),
            ])

        file = BytesIO()
        wb.save(file)
        file.seek(0)

        filename = _sanitize_xlsx_filename(f"{collection.name or 'collection'}_cases.xlsx")
        return await send_file(
            file,
            as_attachment=True,
            attachment_filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return server_error_response(e)


@manager.route('/collection/<collection_id>/cases', methods=['GET'])  # noqa: F821
@login_required
async def get_test_cases(collection_id):
    """Get all test cases for a collection"""
    try:
        if not _has_collection_access(collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 20))
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 20
        if page_size > 200:
            page_size = 200

        result = EvaluationService.list_test_cases(
            collection_id=collection_id, page=page, page_size=page_size
        )
        return get_json_result(data={
            "cases": result.get("cases", []),
            "total": result.get("total", 0),
            "page": page,
            "page_size": page_size,
        })
    except Exception as e:
        return server_error_response(e)


@manager.route('/case/<case_id>', methods=['DELETE'])  # noqa: F821
@login_required
async def delete_test_case(case_id):
    """Delete a test case"""
    try:
        case = EvaluationCase.get_or_none(EvaluationCase.id == case_id)
        if not case:
            return get_data_error_result(message="Test case not found")
        if not _has_collection_access(case.collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.delete_test_case(case_id)

        if not success:
            return get_data_error_result(message="Failed to delete test case")

        return get_json_result(data={"case_id": case_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/case/<case_id>', methods=['PUT'])  # noqa: F821
@login_required
async def update_test_case(case_id):
    """
    Update a test case.

    Request body:
    {
        "variable": {
            "question": "Test question",
            "reference_answer": "Optional ground truth answer"
        },
        "relevant_doc_ids": ["doc_id1", "doc_id2"],
        "relevant_kb_ids": ["kb_id1", "kb_id2"],
        "metadata": {"key": "value"}
    }
    """
    try:
        case = EvaluationCase.get_or_none(EvaluationCase.id == case_id)
        if not case:
            return get_data_error_result(message="Test case not found")
        if not _has_collection_access(case.collection_id):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)

        req = await get_request_json()
        update_data = {
            "variable": req.get("variable"),
            "relevant_doc_ids": req.get("relevant_doc_ids"),
            "relevant_kb_ids": req.get("relevant_kb_ids"),
            "metadata": req.get("metadata"),
        }
        update_data = {k: v for k, v in update_data.items() if v is not None}
        if not update_data:
            return get_data_error_result(message="No valid fields to update")

        success = EvaluationService.update_test_case(case_id, **update_data)
        if not success:
            return get_data_error_result(message="Failed to update test case")

        return get_json_result(data={"case_id": case_id})
    except Exception as e:
        return server_error_response(e)


# ==================== Evaluation Execution ====================
@manager.route('/run/<run_id>/start', methods=['POST'])  # noqa: F821
@login_required
async def start_evaluation(run_id: str):
    try:
        req = await get_request_json()
        case_ids = req.get("case_ids")
        metrics_name = req.get("metrics_name")
        success, task_id = EvaluationService.queue_run_task(
            run_id,
            priority=get_tenant_priority(current_user.id),
            case_ids=case_ids,
            metrics_name=metrics_name,
        )
        if not success:
            raise LookupError("Document not found.")

        return get_json_result(data={"task_id": task_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>', methods=['GET'])  # noqa: F821
@login_required
async def get_evaluation_run(run_id):
    """Get evaluation run details"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(
                message="Evaluation run not found",
                code=RetCode.DATA_ERROR
            )
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)

        ok, task = TaskService.get_by_id(run_id)
        if ok and task:
            run["task"] = task.to_dict()

        return get_json_result(data=run)
    except Exception as e:
        return server_error_response(e)


@manager.route('/run', methods=['PUT'])  # noqa: F821
@login_required
@validate_request("collection_id", "target_type", "target_id", "name")
async def new_run():
    req = await get_request_json() 
    try:
        _, run_id = EvaluationService.create_run_config(
            collection_id=req["collection_id"], 
            target_type=req["target_type"], 
            target_id=req["target_id"],
            user_id=current_user.id, 
            name=req["name"],
            config_snapshot=req.get("config_snapshot")
        )

        return get_json_result(data={"run_id": run_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/results', methods=['GET'])  # noqa: F821
@login_required
async def get_run_results(run_id):
    """Get detailed results for an evaluation run"""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(
                message="Evaluation run not found",
                code=RetCode.DATA_ERROR
            )
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        result = EvaluationService.get_run_results(run_id, page, page_size)

        if not result:
            return get_data_error_result(
                message="Evaluation run not found",
                code=RetCode.DATA_ERROR
            )

        return get_json_result(data=result)
    except Exception as e:
        return server_error_response(e)
    

@manager.route('/run/<run_id>/cancel', methods=['POST'])  # noqa: F821
@login_required
async def cancel_run(run_id):
    """Get detailed results for an evaluation run"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(
                message="Evaluation run not found",
                code=RetCode.DATA_ERROR
            )
        REDIS_CONN.set(f"{run['task_id']}-cancel", "x")
        EvaluationService.cancel_run_task(run_id)
    except Exception as e:
        logging.exception(e)
    return get_json_result(data=True)


@manager.route('/run/list', methods=['GET'])  # noqa: F821
@login_required
async def list_evaluation_runs():
    try:
        target_id = request.args.get("target_id")        
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 20))
        keywords = (request.args.get("keywords", "") or "").strip()
        return get_json_result(data=EvaluationService.list_runs(target_id, page, page_size, keywords))
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>', methods=['DELETE'])  # noqa: F821
@login_required
async def delete_evaluation_run(run_id):
    """Delete an evaluation run"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.delete_run(run_id)
        if not success:
            return get_data_error_result(message="Failed to delete evaluation run")
        return get_json_result(data={"run_id": run_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>', methods=['PUT'])  # noqa: F821
@login_required
async def update_evaluation_run(run_id):
    """
    Update an evaluation run.

    Request body:
    {
        "name": "Optional run name",
        "config_snapshot": {}
    }
    """
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)

        req = await get_request_json()
        update_data = {
            "name": req.get("name"),
            "config_snapshot": req.get("config_snapshot"),
        }
        update_data = {k: v for k, v in update_data.items() if v is not None}
        if not update_data:
            return get_data_error_result(message="No valid fields to update")

        success = EvaluationService.update_run(run_id, **update_data)
        if not success:
            return get_data_error_result(message="Failed to update evaluation run")
        return get_json_result(data={"run_id": run_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/duplicate', methods=['POST'])  # noqa: F821
@login_required
async def duplicate_evaluation_run(run_id):
    """Duplicate an evaluation run"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        name = req.get("name")
        success, new_run_id = EvaluationService.duplicate_run(
            run_id=run_id,
            user_id=current_user.id,
            name=name
        )
        if not success:
            return get_data_error_result(message=new_run_id)
        return get_json_result(data={"run_id": new_run_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/execute', methods=['POST'])  # noqa: F821
@login_required
async def execute_run_case(run_id, case_id):
    """Execute a single case in a run"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.execute_run_case(run_id, case_id, current_user.id)
        if not success:
            return get_data_error_result(message="Failed to execute case")
        return get_json_result(data={"run_id": run_id, "case_id": case_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/metric', methods=['POST'])  # noqa: F821
@login_required
@validate_request("metric_name")
async def run_metric_for_case(run_id, case_id):
    """Compute a single metric for a case"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        metric_name = req.get("metric_name")
        success = EvaluationService.run_metrics_for_case(
            run_id, case_id, current_user.id, [metric_name]
        )
        if not success:
            return get_data_error_result(message="Failed to compute metric")
        return get_json_result(data={"run_id": run_id, "case_id": case_id, "metric": metric_name})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/metrics', methods=['POST'])  # noqa: F821
@login_required
async def run_metrics_for_cases(run_id):
    """
    Compute all metrics for all cases in a run.
    """
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success_count, failure_count = EvaluationService.run_metrics_for_cases(
            run_id, current_user.id
        )
        total = len(EvaluationService.get_test_cases(run["collection_id"]))
        return get_json_result(data={
            "run_id": run_id,
            "success_count": success_count,
            "failure_count": failure_count,
            "total": total
        })
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/metric', methods=['POST'])  # noqa: F821
@login_required
@validate_request("metric_name")
async def run_metric_for_cases(run_id):
    """
    Compute a single metric for all cases in a run.

    """
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        metric_name = req.get("metric_name")
        success_count, failure_count = EvaluationService.run_metric_for_cases(
            run_id, metric_name, current_user.id
        )
        total = len(EvaluationService.get_test_cases(run["collection_id"]))
        return get_json_result(data={
            "run_id": run_id,
            "metric": metric_name,
            "success_count": success_count,
            "failure_count": failure_count,
            "total": total
        })
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/metrics', methods=['POST'])  # noqa: F821
@login_required
async def run_metrics_for_case(run_id, case_id):
    """Compute all metrics for a case"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.run_metrics_for_case(
            run_id, case_id, current_user.id
        )
        if not success:
            return get_data_error_result(message="Failed to compute metrics")
        return get_json_result(data={"run_id": run_id, "case_id": case_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/result', methods=['DELETE'])  # noqa: F821
@login_required
async def clear_result(run_id, case_id):
    """Clear a result without deleting the record"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.clear_result(run_id, case_id)
        if not success:
            return get_data_error_result(message="Failed to clear result")
        return get_json_result(data={"run_id": run_id, "case_id": case_id})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/result/metric', methods=['DELETE'])  # noqa: F821
@login_required
@validate_request("metric_name")
async def clear_result_metric(run_id, case_id):
    """Clear a single metric from result"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        req = await get_request_json()
        metric_name = req.get("metric_name")
        success = EvaluationService.clear_result_metric(run_id, case_id, metric_name)
        if not success:
            return get_data_error_result(message="Failed to clear metric")
        return get_json_result(data={"run_id": run_id, "case_id": case_id, "metric": metric_name})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/case/<case_id>/result/answer', methods=['DELETE'])  # noqa: F821
@login_required
async def clear_result_generated_answer(run_id, case_id):
    """Clear generated answer from result"""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(message="Evaluation run not found")
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)
        success = EvaluationService.clear_result_generated_answer(run_id, case_id)
        if not success:
            return get_data_error_result(message="Failed to clear generated answer")
        return get_json_result(data={"run_id": run_id, "case_id": case_id})
    except Exception as e:
        return server_error_response(e)


# ==================== Analysis & Recommendations ====================

@manager.route('/run/<run_id>/recommendations', methods=['GET'])  # noqa: F821
@login_required
async def get_recommendations(run_id):
    """Get configuration recommendations based on evaluation results"""
    try:
        return get_json_result(data={"recommendations": []})
    except Exception as e:
        return server_error_response(e)


@manager.route('/compare', methods=['POST'])  # noqa: F821
@login_required
@validate_request("run_ids")
async def compare_runs():
    """
    Compare multiple evaluation runs.

    Request body:
    {
        "run_ids": ["run_id1", "run_id2", "run_id3"]
    }
    """
    try:
        req = await get_request_json()
        run_ids = req.get("run_ids", [])

        if not run_ids or not isinstance(run_ids, list) or len(run_ids) < 2:
            return get_data_error_result(
                message="run_ids must be a list with at least 2 run IDs"
            )

        return get_json_result(data={"comparison": {}})
    except Exception as e:
        return server_error_response(e)


@manager.route('/run/<run_id>/export', methods=['GET', 'POST'])  # noqa: F821
@login_required
async def export_results(run_id):
    """Export evaluation results as an Excel file (.xlsx)."""
    try:
        run = EvaluationService.get_run(run_id)
        if not run:
            return get_data_error_result(
                message="Evaluation run not found",
                code=RetCode.DATA_ERROR
            )
        if not _has_collection_access(run["collection_id"]):
            return get_json_result(
                data=False, message="Collection not configured or access denied",
                code=RetCode.OPERATING_ERROR)

        dumps = json.dumps
        safe_text = _xlsx_safe_text
        metrics_summary = run.get("metrics_summary")
        metric_keys = (
            sorted(str(k) for k in metrics_summary.keys())
            if isinstance(metrics_summary, dict) and metrics_summary
            else []
        )

        def _to_cell(v):
            if v is None:
                return ""
            if isinstance(v, (int, float, bool)):
                return v
            if isinstance(v, str):
                return safe_text(v)
            try:
                return safe_text(dumps(v, ensure_ascii=False, separators=(",", ":")))
            except Exception:
                return safe_text(v)

        wb = Workbook(write_only=True)

        ws_run = wb.create_sheet("run")
        ws_run_append = ws_run.append
        ws_run_append(["key", "value"])
        for k, v in (run or {}).items():
            ws_run_append([safe_text(k), _to_cell(v)])

        ws = wb.create_sheet("results")
        ws_append = ws.append
        ws_append([
            "question",
            "reference_answer",
            "generated_answer",
            "execution_time",
            "token_usage",
            "retrieved_chunk_count",
            *metric_keys,
        ])

        with DB.connection_context():
            query = (
                EvaluationResult
                .select(
                    EvaluationCase.variable,
                    EvaluationResult.execution_time,
                    EvaluationResult.token_usage,
                    EvaluationResult.generated_answer,
                    EvaluationResult.retrieved_chunks,
                    EvaluationResult.metrics,
                )
                .join(EvaluationCase, on=(EvaluationResult.case_id == EvaluationCase.id))
                .where(EvaluationResult.run_id == run_id)
                .order_by(EvaluationResult.id)
                .dicts()
            )

            to_cell = _to_cell
            for row in query.iterator():
                variable = row.get("variable") or {}
                metrics = row.get("metrics") or {}
                if not isinstance(metrics, dict):
                    metrics = {}

                retrieved_chunks = row.get("retrieved_chunks") or []
                if isinstance(retrieved_chunks, list) and retrieved_chunks:
                    doc_ids = {
                        c.get("doc_id")
                        for c in retrieved_chunks
                        if isinstance(c, dict) and c.get("doc_id")
                    }
                    retrieved_doc_ids = ",".join(sorted(doc_ids)) if doc_ids else ""
                    retrieved_chunk_count = len(retrieved_chunks)
                else:
                    retrieved_doc_ids = ""
                    retrieved_chunk_count = 0

                ws_append([
                    to_cell(variable.get("question", "")),
                    to_cell(variable.get("reference_answer", "")),
                    to_cell(row.get("generated_answer")),
                    row.get("execution_time", ""),
                    to_cell(row.get("token_usage")),
                    retrieved_chunk_count,
                    to_cell(retrieved_doc_ids),
                    *[to_cell(metrics.get(k, "")) for k in metric_keys],
                ])

        file = BytesIO()
        wb.save(file)
        file.seek(0)

        filename = _sanitize_xlsx_filename(f"{run.get('name') or run_id}_results.xlsx")
        return await send_file(
            file,
            as_attachment=True,
            attachment_filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return server_error_response(e)


# ==================== Real-time Evaluation ====================

@manager.route('/evaluate_single', methods=['POST'])  # noqa: F821
@login_required
@validate_request("question", "dialog_id")
async def evaluate_single():
    """
    Evaluate a single question-answer pair in real-time.

    Request body:
    {
        "question": "Test question",
        "dialog_id": "dialog_id",
        "reference_answer": "Optional ground truth"
    }
    """
    try:
        # req = await get_request_json()  # TODO: Use for single evaluation implementation

        # TODO: Implement single evaluation
        # This would execute the RAG pipeline and return metrics immediately

        return get_json_result(data={
            "answer": "",
            "metrics": {},
            "retrieved_chunks": []
        })
    except Exception as e:
        return server_error_response(e)
